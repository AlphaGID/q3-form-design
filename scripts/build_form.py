"""
Builds the HH2026 XLSForm from plain Python lists.
Re-run this whole script every time the form changes; don't hand-edit the .xlsx.
"""
import csv
import openpyxl

HEADER = ["type", "name", "label", "hint", "required", "relevant",
          "constraint", "constraint_message", "calculation", "appearance",
          "choice_filter", "default", "readonly", "repeat_count"]

def row(type_, name, label, hint="", required="", relevant="", constraint="",
        constraint_message="", calculation="", appearance="", choice_filter="",
        default="", readonly="", repeat_count=""):
    return [type_, name, label, hint, required, relevant, constraint,
            constraint_message, calculation, appearance, choice_filter,
            default, readonly, repeat_count]

wb = openpyxl.Workbook()

# --- settings ---
settings_sheet = wb.active
settings_sheet.title = "settings"
settings_sheet.append(["form_title", "form_id", "version", "default_language", "style"])
settings_sheet.append(["Integrated Child Health and AMR Survey 2026", "hh2026_v1", "2026070200", "en", "pages"])

# --- survey ---
survey_sheet = wb.create_sheet("survey")
survey_sheet.append(HEADER)

survey_rows = [
    row("begin group", "section1", "Section 1: Household identification"),

    row("select_one lga_list", "lga", "1.02 Local Government Area", required="yes"),

    row("select_one_from_file wards.csv", "ward", "1.03 Ward",
        required="yes", choice_filter="lga_code=${lga}"),

    row("select_one_from_file settlements.csv", "settlement", "1.04 Settlement",
        required="yes", choice_filter="ward_code=${ward}"),

    row("select_one yesno", "settlement_known_locally",
        "1.05 Is the settlement known locally by a different name?", required="yes"),
    row("text", "settlement_local_name", "Name used locally",
        relevant="${settlement_known_locally}='1'",
        required="${settlement_known_locally}='1'"),

    row("integer", "structure_number", "1.06 Structure number painted on the dwelling",
        required="yes", constraint=". >= 1 and . <= 999",
        constraint_message="Enter a number between 1 and 999"),

    row("integer", "hh_serial_number", "1.07 Household serial number within the settlement",
        required="yes", constraint=". >= 1 and . <= 999",
        constraint_message="Enter a number between 1 and 999"),

    row("select_one_from_file staff_roster.csv", "enumerator_code", "1.08 Enumerator code",
        required="yes", choice_filter="role='Enumerator'"),

    row("calculate", "team_code_calc", "",
        calculation="pulldata('staff_roster','team_code','name',${enumerator_code})"),
    row("text", "team_code_display", "1.09 Team code",
        calculation="${team_code_calc}", readonly="yes"),

    row("date", "visit_date", "1.10 Date of visit", required="yes",
        constraint=". >= date('2026-06-01') and . <= date('2026-06-30')",
        constraint_message="Date must be within the fieldwork period, 1-30 June 2026"),

    row("geopoint", "gps_reading",
        "1.11 Record the GPS reading taken at the entrance to the dwelling.", required="yes"),
    row("calculate", "gps_lat", "", calculation="substring-before(${gps_reading},' ')"),
    row("calculate", "gps_lon", "",
        calculation="substring-before(substring-after(${gps_reading},' '),' ')"),
    row("calculate", "settlement_lat_calc", "",
        calculation="pulldata('settlements','latitude','name',${settlement})"),
    row("calculate", "settlement_lon_calc", "",
        calculation="pulldata('settlements','longitude','name',${settlement})"),
    row("note", "gps_plausibility_flag",
        "This GPS reading looks far from the registered location of the selected settlement. Please confirm this is the correct dwelling before continuing.",
        relevant="(${gps_reading}!='') and ((abs(number(${gps_lat}) - number(${settlement_lat_calc})) > 0.05) or (abs(number(${gps_lon}) - number(${settlement_lon_calc})) > 0.05))"),

    row("select_one visited_previous_list", "prev_round_visited",
        "1.12 Was this household visited during the October 2025 round?", required="yes"),
    row("text", "prev_household_id",
        "1.13 Record the household identifier allocated in the October 2025 round.",
        relevant="${prev_round_visited}='1'",
        constraint="pulldata('previous_round_households','household_id','household_id',.) != ''",
        constraint_message="This household ID was not found in the previous round register. Please check and re-enter."),

    row("select_one visit_result_list", "result_of_visit", "1.14 Result of visit", required="yes"),

    row("end group", "", ""),

    row("begin group", "section2", "Section 2: Consent", relevant="${result_of_visit}='1'"),
    row("select_one yesno", "consent_read_aloud",
        "2.01 Consent statement read aloud to the respondent in full?", required="yes"),
    row("select_one consent_list", "consent_given",
        "2.02 Does the respondent consent to the household interview?", required="yes"),
    row("select_one relationship_list", "respondent_relationship",
        "2.03 Relationship of the respondent to the head of household",
        required="yes", relevant="${consent_given}='1'"),
    row("end group", "", ""),

    row("begin group", "section3", "Section 3: Household roster",
        relevant="${result_of_visit}='1' and ${consent_given}='1'"),

    row("integer", "household_size", "3.01 How many people usually live in this household?",
        required="yes", constraint=". >= 1 and . <= 30",
        constraint_message="Enter a number between 1 and 30"),

    row("begin repeat", "roster", "Household roster", repeat_count="${household_size}"),
    row("text", "name_initials", "(2) Name or initials", required="yes"),
    row("select_one relationship_list", "relationship_to_head", "(3) Relationship to head", required="yes"),
    row("select_one sex_list", "sex", "(4) Sex", required="yes"),
    row("select_one yesno", "under5", "Is this person under 5 years old?", required="yes"),
    row("integer", "age_years", "(5) Age in completed years",
        relevant="${under5}='2'", required="${under5}='2'",
        constraint=". >= 5 and . <= 120", constraint_message="Enter age in completed years, 5 or older"),
    row("integer", "age_months", "(6) Age in completed months (under 5 only)",
        relevant="${under5}='1'", required="${under5}='1'",
        constraint=". >= 0 and . <= 59", constraint_message="Enter age in completed months, 0 to 59"),
    row("calculate", "eligible_for_section4", "",
        calculation="if(${under5}='1' and ${age_months} >= 9 and ${age_months} <= 59, 1, 0)"),

    row("begin group", "section4_module", "Section 4: Child module",
        relevant="${eligible_for_section4}='1'"),

    row("calculate", "roster_line_number", "", calculation="position(..)"),
    row("note", "s4_child_ref",
        "This module is for: ${name_initials}, ${age_months} months, line ${roster_line_number}"),

    row("select_one measured_status_list", "weight_status", "4.05 Weight of the child", required="yes"),
    row("decimal", "weight_kg", "Weight (kg)",
        relevant="${weight_status}='1'", required="${weight_status}='1'",
        constraint=". >= 2.0 and . <= 30.0",
        constraint_message="Enter weight in kg, 2.0 to 30.0"),

    row("select_one measured_status_list", "height_status", "4.06 Length or height of the child", required="yes"),
    row("decimal", "height_cm", "Length or height (cm)",
        relevant="${height_status}='1'", required="${height_status}='1'",
        constraint=". >= 60.0 and . <= 120.0",
        constraint_message="Enter length/height in cm, 60.0 to 120.0"),
    row("select_one position_list", "measurement_position", "4.07 Position in which the child was measured",
        relevant="${height_status}='1'", required="${height_status}='1'"),

    row("select_one card_seen_list", "card_seen",
        "4.08 May I see the child's vaccination card or health record?", required="yes"),
    row("select_one yesno", "measles_on_card",
        "4.09 Copy from the card: is a measles dose recorded?",
        relevant="${card_seen}='1'", required="${card_seen}='1'"),
    row("select_one yesnodk_list", "measles_ever",
        "4.10 Has this child ever received a measles vaccination?",
        relevant="${card_seen}='2'", required="${card_seen}='2'"),

    row("select_one yesnodk_list", "diarrhoea_14d",
        "4.11 Has this child had diarrhoea in the past 14 days?", required="yes"),

    row("select_one yesnodk_list", "antibiotic_30d",
        "4.12 Has this child taken any antibiotic medicine in the past 30 days?", required="yes"),
    row("text", "antibiotic_name",
        "4.13/4.14 Which antibiotic was taken? Record the name as reported by the caregiver.",
        hint="OPEN ITEM: no controlled medicine list was supplied with the data pack (see defects log D01). Recorded as free text pending that list.",
        relevant="${antibiotic_30d}='1'", required="${antibiotic_30d}='1'"),
    row("select_one yesnodk_list", "antibiotic_no_prescription",
        "4.15 Was the medicine obtained without a prescription from a health worker?",
        relevant="${antibiotic_30d}='1'", required="${antibiotic_30d}='1'"),
    row("select_one photo_taken_list", "antibiotic_photo_taken",
        "4.16 Was a photograph of the medicine packaging taken?",
        relevant="${antibiotic_30d}='1'", required="${antibiotic_30d}='1'"),

    row("end group", "", ""),

    row("begin group", "section5_module", "Section 5: Specimen collection",
        relevant="${eligible_for_section4}='1'"),

    row("select_one yesno", "age12plus",
        "5.01 Is the child aged 12 completed months or older?", required="yes"),

    row("select_one yesno", "specimen_obtained",
        "5.02 Was a stool specimen obtained from this child?",
        relevant="${age12plus}='1'", required="${age12plus}='1'"),

    row("text", "specimen_number",
        "5.03 Specimen label number (6 digits printed after the BSN prefix)",
        relevant="${specimen_obtained}='1'", required="${specimen_obtained}='1'",
        calculation="",
        constraint="regex(.,'^[0-9]{6}$') and number(.) >= number(${range_start_calc}) and number(.) <= number(${range_end_calc})",
        constraint_message="Must be exactly 6 digits and within your team's issued label range (${range_start_calc}-${range_end_calc})"),

    row("calculate", "range_start_calc", "",
        calculation="pulldata('specimen_label_allocation','range_start','team_code',${team_code_calc})"),
    row("calculate", "range_end_calc", "",
        calculation="pulldata('specimen_label_allocation','range_end','team_code',${team_code_calc})"),

    row("calculate", "specimen_digit0", "",
        calculation="if(string-length(${specimen_number})=6, number(substr(${specimen_number},0,1)), 0)"),
    row("calculate", "specimen_digit1", "",
        calculation="if(string-length(${specimen_number})=6, number(substr(${specimen_number},1,2)), 0)"),
    row("calculate", "specimen_digit2", "",
        calculation="if(string-length(${specimen_number})=6, number(substr(${specimen_number},2,3)), 0)"),
    row("calculate", "specimen_digit3", "",
        calculation="if(string-length(${specimen_number})=6, number(substr(${specimen_number},3,4)), 0)"),
    row("calculate", "specimen_digit4", "",
        calculation="if(string-length(${specimen_number})=6, number(substr(${specimen_number},4,5)), 0)"),
    row("calculate", "specimen_digit5", "",
        calculation="if(string-length(${specimen_number})=6, number(substr(${specimen_number},5,6)), 0)"),

    row("calculate", "specimen_weighted_sum", "",
        calculation="(${specimen_digit0}*7)+(${specimen_digit1}*6)+(${specimen_digit2}*5)+(${specimen_digit3}*4)+(${specimen_digit4}*3)+(${specimen_digit5}*2)"),
    row("calculate", "specimen_remainder", "", calculation="${specimen_weighted_sum} mod 11"),
    row("calculate", "specimen_computed_check_digit", "",
        calculation="if(${specimen_remainder}=0,'0',if(${specimen_remainder}=1,'X',string(11 - ${specimen_remainder})))"),

    row("text", "specimen_check_digit", "Check digit (single digit 0-9, or X)",
        relevant="${specimen_obtained}='1'", required="${specimen_obtained}='1'",
        constraint=". = ${specimen_computed_check_digit}",
        constraint_message="Check digit does not match this label number. Please re-check both and re-enter."),

    row("time", "specimen_coldbox_time", "5.04 Time the specimen was placed in the cold box",
        relevant="${specimen_obtained}='1'", required="${specimen_obtained}='1'"),

    row("decimal", "specimen_coldbox_temp", "5.05 Temperature shown on the cold box thermometer at that time",
        relevant="${specimen_obtained}='1'", required="${specimen_obtained}='1'",
        constraint=". >= 2.0 and . <= 8.0",
        constraint_message="Enter a temperature between 2.0 and 8.0 degrees C"),

    row("select_one reason_no_specimen_list", "reason_no_specimen", "5.06 Reason no specimen was obtained",
        relevant="${specimen_obtained}='2'", required="${specimen_obtained}='2'"),
    row("text", "reason_no_specimen_other", "5.07 If code 96, specify",
        relevant="${reason_no_specimen}='96'", required="${reason_no_specimen}='96'"),

    row("end group", "", ""),

    row("end repeat", "", ""),

    row("calculate", "eligible_children_count", "", calculation="sum(../roster/eligible_for_section4)"),
    row("text", "eligible_children_display",
        "3.02 Number of children aged 9 to 59 completed months (calculated automatically from the roster)",
        calculation="${eligible_children_count}", readonly="yes"),

    row("end group", "", ""),

    row("begin group", "section6", "Section 6: Household environment",
        relevant="${result_of_visit}='1' and ${consent_given}='1'"),

    row("select_one water_source_list", "water_source",
        "6.01 What is the main source of drinking water for members of this household?", required="yes"),
    row("select_one toilet_list", "toilet_type",
        "6.02 What kind of toilet facility do members of this household usually use?", required="yes"),
    row("select_one yesno", "keeps_animals",
        "6.03 Does this household keep poultry or livestock inside the compound?", required="yes"),
    row("select_one yesnodk_list", "animal_antibiotics_12m",
        "6.04 Have any antibiotic medicines been given to these animals in the past 12 months?",
        relevant="${keeps_animals}='1'", required="${keeps_animals}='1'"),
    row("select_one handwashing_list", "handwashing_station",
        "6.05 Observe: is there a handwashing station with both soap and water available?", required="yes"),
    row("select_one yesnodk_list", "hh_diarrhoea_2w",
        "6.06 Has any member of this household had diarrhoea in the past two weeks?", required="yes"),
    row("select_multiple assets_list", "household_assets",
        "6.07 Which of the following does this household own?", required="yes",
        constraint="not(selected(., 'H')) or count-selected(.) = 1",
        constraint_message="'None of these' cannot be selected together with any other option"),

    row("end group", "", ""),

    row("begin group", "section7", "Section 7: Close-out and supervisor review",
        relevant="${result_of_visit}='1' and ${consent_given}='1'"),

    row("time", "interview_end_time", "7.01 Time the interview ended", required="yes"),
    row("text", "office_observation",
        "7.02 Record any observation that may help the office interpret this form."),
    row("text", "enumerator_signature", "7.03 Enumerator signature", required="yes"),
    row("date", "enumerator_signature_date", "Enumerator signature date", required="yes"),

    row("select_one_from_file staff_roster.csv", "supervisor_code", "7.04 Supervisor code",
        required="yes", choice_filter="role='Supervisor'"),
    row("select_one supervisor_decision_list", "supervisor_decision",
        "7.05 Supervisor decision on this form", required="yes"),
    row("text", "supervisor_signature", "7.06 Supervisor signature", required="yes"),
    row("date", "supervisor_signature_date", "Supervisor signature date", required="yes"),

    row("end group", "", ""),
]

for r in survey_rows:
    survey_sheet.append(r)

# --- choices ---
choices_sheet = wb.create_sheet("choices")
choices_sheet.append(["list_name", "name", "label"])

# LGA list read from the actual data file, not retyped
with open("data/reference/lgas.csv") as f:
    reader = csv.DictReader(f)
    for lga in reader:
        choices_sheet.append(["lga_list", lga["name"], lga["label"]])

choices_sheet.append(["yesno", "1", "Yes"])
choices_sheet.append(["yesno", "2", "No"])

choices_sheet.append(["visited_previous_list", "1", "Yes"])
choices_sheet.append(["visited_previous_list", "2", "No"])
choices_sheet.append(["visited_previous_list", "8", "Do not know"])

choices_sheet.append(["visit_result_list", "1", "Completed"])
choices_sheet.append(["visit_result_list", "2", "Refused"])
choices_sheet.append(["visit_result_list", "3", "No competent adult after three visits"])
choices_sheet.append(["visit_result_list", "4", "Dwelling vacant or demolished"])

choices_sheet.append(["consent_list", "1", "Consent given"])
choices_sheet.append(["consent_list", "2", "Consent refused"])

choices_sheet.append(["relationship_list", "1", "Head"])
choices_sheet.append(["relationship_list", "2", "Spouse"])
choices_sheet.append(["relationship_list", "3", "Son or daughter"])
choices_sheet.append(["relationship_list", "4", "Parent"])
choices_sheet.append(["relationship_list", "5", "Other relative"])
choices_sheet.append(["relationship_list", "6", "Not related"])

choices_sheet.append(["sex_list", "1", "Male"])
choices_sheet.append(["sex_list", "2", "Female"])

choices_sheet.append(["yesnodk_list", "1", "Yes"])
choices_sheet.append(["yesnodk_list", "2", "No"])
choices_sheet.append(["yesnodk_list", "8", "Do not know"])

choices_sheet.append(["measured_status_list", "1", "Measured"])
choices_sheet.append(["measured_status_list", "2", "Not measured"])

choices_sheet.append(["card_seen_list", "1", "Card seen"])
choices_sheet.append(["card_seen_list", "2", "No card seen"])

choices_sheet.append(["position_list", "1", "Recumbent length"])
choices_sheet.append(["position_list", "2", "Standing height"])

choices_sheet.append(["photo_taken_list", "1", "Yes"])
choices_sheet.append(["photo_taken_list", "2", "No, not available"])
choices_sheet.append(["photo_taken_list", "3", "Caregiver declined"])

choices_sheet.append(["reason_no_specimen_list", "1", "Caregiver refused"])
choices_sheet.append(["reason_no_specimen_list", "2", "Child absent"])
choices_sheet.append(["reason_no_specimen_list", "3", "Unable to produce"])
choices_sheet.append(["reason_no_specimen_list", "4", "Container spoiled"])
choices_sheet.append(["reason_no_specimen_list", "96", "Other"])

choices_sheet.append(["water_source_list", "1", "Piped into dwelling"])
choices_sheet.append(["water_source_list", "2", "Piped into compound"])
choices_sheet.append(["water_source_list", "3", "Public tap or standpipe"])
choices_sheet.append(["water_source_list", "4", "Tube well or borehole"])
choices_sheet.append(["water_source_list", "5", "Protected dug well"])
choices_sheet.append(["water_source_list", "6", "Unprotected dug well"])
choices_sheet.append(["water_source_list", "7", "Protected spring"])
choices_sheet.append(["water_source_list", "8", "Unprotected spring"])
choices_sheet.append(["water_source_list", "9", "Rainwater"])
choices_sheet.append(["water_source_list", "10", "Tanker or cart"])
choices_sheet.append(["water_source_list", "11", "Surface water"])

choices_sheet.append(["toilet_list", "1", "Flush to sewer"])
choices_sheet.append(["toilet_list", "2", "Flush to septic tank"])
choices_sheet.append(["toilet_list", "3", "Flush to pit latrine"])
choices_sheet.append(["toilet_list", "4", "Ventilated improved pit"])
choices_sheet.append(["toilet_list", "5", "Pit latrine with slab"])
choices_sheet.append(["toilet_list", "6", "Pit latrine without slab"])
choices_sheet.append(["toilet_list", "7", "Composting toilet"])
choices_sheet.append(["toilet_list", "8", "Bucket"])
choices_sheet.append(["toilet_list", "9", "No facility or bush"])

choices_sheet.append(["handwashing_list", "1", "Observed, soap and water"])
choices_sheet.append(["handwashing_list", "2", "Reported only, not observed"])
choices_sheet.append(["handwashing_list", "3", "Not present"])

choices_sheet.append(["assets_list", "A", "Radio"])
choices_sheet.append(["assets_list", "B", "Television"])
choices_sheet.append(["assets_list", "C", "Mobile telephone"])
choices_sheet.append(["assets_list", "D", "Bicycle"])
choices_sheet.append(["assets_list", "E", "Motorcycle"])
choices_sheet.append(["assets_list", "F", "Car or truck"])
choices_sheet.append(["assets_list", "G", "Refrigerator"])
choices_sheet.append(["assets_list", "H", "None of these"])

choices_sheet.append(["supervisor_decision_list", "1", "Accept"])
choices_sheet.append(["supervisor_decision_list", "2", "Return for correction"])
choices_sheet.append(["supervisor_decision_list", "3", "Void"])

wb.save("form/HH2026v1.xlsx")
print("form/HH2026v1.xlsx written with Section 1")
